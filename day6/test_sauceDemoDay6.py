import pytest
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from selenium.webdriver. support.wait import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from constants import globalConstants as gc


class Test_SauceDemoClass:

    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.get(gc.mainURL)
        self.driver.maximize_window()

    def teardown_method(self):
        self.driver.quit()

    def getInvalidLoginDatas():
        excelFile = openpyxl.load_workbook("day6/data/invalidLoginDatas.xlsx")
        selectedSheet = excelFile["Sayfa1"]
        data = []

        totalRows = selectedSheet.max_row

        for i in range(2, totalRows+1):
            username = selectedSheet.cell(i, 1).value
            password = selectedSheet.cell(i, 2).value
            dataTuple = (username, password)
            data.append(dataTuple)

        return data

    @pytest.mark.parametrize("username,password", getInvalidLoginDatas())
    def test_invalidLogins(self, username, password):
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
            (By.CSS_SELECTOR, "*[data-test=\"username\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
            (By.CSS_SELECTOR, "*[data-test=\"password\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
            (By.CSS_SELECTOR, "*[data-test=\"login-button\"]")))

        usernameInput = self.driver.find_element(
            By.CSS_SELECTOR, "*[data-test=\"username\"]")
        usernameInput.click()
        self.driver.find_element(
            By.CSS_SELECTOR, "*[data-test=\"username\"]").send_keys("1")
        passwordInput = self.driver.find_element(
            By.CSS_SELECTOR, "*[data-test=\"password\"]")
        passwordInput.click()
        self.driver.find_element(
            By.CSS_SELECTOR, "*[data-test=\"password\"]").send_keys("1")
        loginButton = self.driver.find_element(
            By.CSS_SELECTOR, "*[data-test=\"login-button\"]")
        loginButton.click()
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located(
            (By.CSS_SELECTOR, "*[data-test=\"error\"]")))
        assert self.driver.find_element(
        By.CSS_SELECTOR, "*[data-test=\"error\"]").text == "Epic sadface: Username and password do not match any user in this service"

    def test_testvalidLogin(self):
        
        

        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"username\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"password\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"login-button\"]")))
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").send_keys(gc.validUsername)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"password\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"password\"]").send_keys(gc.validPassword)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"login-button\"]").click()

        assert self.driver.current_url == gc.expectedURLThenLogIn
    
    def test_testlogOutFromBurgerMenu(self):
    
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"username\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"password\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"login-button\"]")))
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").send_keys(gc.validUsername)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"password\"]").send_keys(gc.validPassword)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"login-button\"]").click()
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, "react-burger-menu-btn")))
        self.driver.find_element(By.ID, "react-burger-menu-btn").click()
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.ID, "logout_sidebar_link")))
        self.driver.find_element(By.ID, "logout_sidebar_link").click()

        assert self.driver.current_url == gc.mainURL

    def test_sortNameAToZ(self):
    
        filterAToZXPath = "//*[@id='header_container']/div[2]/div/span/select/option[1]"
        productNameDivsXPath = "//div[@class='inventory_item_name']"

    
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"username\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"password\"]")))
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"login-button\"]")))
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"username\"]").send_keys(gc.validUsername)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"password\"]").send_keys(gc.validPassword)
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"login-button\"]").click()
        self.driver.find_element(By.CSS_SELECTOR, "*[data-test=\"product_sort_container\"]").click()
        WebDriverWait(self.driver, 5).until(ec.visibility_of_element_located((By.CSS_SELECTOR, "*[data-test=\"product_sort_container\"]")))
        sortMenu = self.getElementByLocator(By.CSS_SELECTOR,"*[data-test=\"product_sort_container\"]")
        
        

    